import pandas as pd
import os
import openpyxl as px

# 50個のExcelファイルが保存されているフォルダへのパス
folder_path = 'occurrenceDataSampleFolder'


k=3



cclist = [
    'AD', 'AE', 'AF', 'AG', 'AI', 'AL', 'AM', 'AO', 'AQ', 'AR', 'AS', 'AT',
    'AU', 'AW', 'AX', 'AZ', 'BA', 'BB', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI',
    'BJ', 'BL', 'BM', 'BN', 'BO', 'BQ', 'BR', 'BS', 'BT', 'BV', 'BW', 'BY',
    'BZ', 'CA', 'CC', 'CD', 'CF', 'CG', 'CH', 'CI', 'CK', 'CL', 'CM', 'CN',
    'CO', 'CR', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DE', 'DJ', 'DK', 'DM',
    'DO', 'DZ', 'EC', 'EE', 'EG', 'EH', 'ER', 'ES', 'ET', 'FI', 'FJ', 'FK',
    'FM', 'FO', 'FR', 'GA', 'GB', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GL',
    'GM', 'GN', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GW', 'GY', 'HK', 'HM',
    'HN', 'HR', 'HT', 'HU', 'ID', 'IE', 'IL', 'IM', 'IN', 'IO', 'IQ', 'IR',
    'IS', 'IT', 'JE', 'JM', 'JO', 'JP', 'KE', 'KG', 'KH', 'KI', 'KM', 'KN',
    'KP', 'KR', 'KW', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LI', 'LK', 'LR', 'LS',
    'LT', 'LU', 'LV', 'LY', 'MA', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH', 'MK',
    'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR', 'MS', 'MT', 'MU', 'MV', 'MW',
    'MX', 'MY', 'MZ', 'NA', 'NC', 'NE', 'NF', 'NG', 'NI', 'NL', 'NO', 'NP',
    'NR', 'NU', 'NZ', 'OM', 'PA', 'PE', 'PF', 'PG', 'PH', 'PK', 'PL', 'PM',
    'PN', 'PR', 'PS', 'PT', 'PW', 'PY', 'QA', 'RE', 'RO', 'RS', 'RU', 'RW',
    'SA', 'SB', 'SC', 'SD', 'SE', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL', 'SM',
    'SN', 'SO', 'SR', 'SS', 'ST', 'SV', 'SX', 'SY', 'SZ', 'TC', 'TD', 'TF',
    'TG', 'TH', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TR', 'TT', 'TV', 'TW',
    'TZ', 'UA', 'UG', 'UM', 'US', 'UY', 'UZ', 'VA', 'VC', 'VE', 'VG', 'VI',
    'VN', 'VU', 'WF', 'WS', 'YE', 'YT', 'ZA', 'ZM', 'ZW' ]

newWb =px.Workbook()
ws2 = newWb.active

# 1行目にCountryCodeを書き込む
ws2.cell(row=1, column=1).value = 'CountryCode'
for i, cc in enumerate(cclist, start=2):
    ws2.cell(row=i, column=1).value = cc

cclist3letter = [
    "AND", "ARE", "AFG", "ATG", "AIA", "ALB", "ARM", "AGO", "ATA", "ARG", "ASM",
    "AUT", "AUS", "ABW", "ALA", "AZE", "BIH", "BRB", "BGD", "BEL", "BFA", "BGR",
    "BHR", "BDI", "BEN", "BLM", "BMU", "BRN", "BOL", "BES", "BRA", "BHS", "BTN",
    "BVT", "BWA", "BLR", "BLZ", "CAN", "CCK", "COD", "CAF", "COG", "CHE", "CIV",
    "COK", "CHL", "CMR", "CHN", "COL", "CRI", "CUB", "CPV", "CUW", "CXR", "CYP",
    "CZE", "DEU", "DJI", "DNK", "DMA", "DOM", "DZA", "ECU", "EST", "EGY", "ESH",
    "ERI", "ESP", "ETH", "FIN", "FJI", "FLK", "FSM", "FRO", "FRA", "GAB", "GBR",
    "GRD", "GEO", "GUF", "GGY", "GHA", "GIB", "GRL", "GMB", "GIN", "GLP", "GNQ",
    "GRC", "SGS", "GTM", "GUM", "GNB", "GUY", "HKG", "HMD", "HND", "HRV", "HTI",
    "HUN", "IDN", "IRL", "ISR", "IMN", "IND", "IOT", "IRQ", "IRN", "ISL", "ITA",
    "JEY", "JAM", "JOR", "JPN", "KEN", "KGZ", "KHM", "KIR", "COM", "KNA", "PRK",
    "KOR", "KWT", "CYM", "KAZ", "LAO", "LBN", "LCA", "LIE", "LKA", "LBR", "LSO",
    "LTU", "LUX", "LVA", "LBY", "MAR", "MCO", "MDA", "MNE", "MAF", "MDG", "MHL",
    "MKD", "MLI", "MMR", "MNG", "MAC", "MNP", "MTQ", "MRT", "MSR", "MLT", "MUS",
    "MDV", "MWI", "MEX", "MYS", "MOZ", "NAM", "NCL", "NER", "NFK", "NGA", "NIC",
    "NLD", "NOR", "NPL", "NRU", "NIU", "NZL", "OMN", "PAN", "PER", "PYF", "PNG",
    "PHL", "PAK", "POL", "SPM", "PCN", "PRI", "PSE", "PRT", "PLW", "PRY", "QAT",
    "REU", "ROU", "SRB", "RUS", "RWA", "SAU", "SLB", "SYC", "SDN", "SWE", "SGP",
    "SHN", "SVN", "SJM", "SVK", "SLE", "SMR", "SEN", "SOM", "SUR", "SSD", "STP",
    "SLV", "SXM", "SYR", "SWZ", "TCA", "TCD", "ATF", "TGO", "THA", "TJK", "TKL",
    "TLS", "TKM", "TUN", "TON", "TUR", "TTO", "TUV", "TWN", "TZA", "UKR", "UGA",
    "UMI", "USA", "URY", "UZB", "VAT", "VCT", "VEN", "VGB", "VIR", "VNM", "VUT",
    "WLF", "WSM", "YEM", "MYT", "ZAF", "ZMB", "ZWE"
]

# 2行目にCountryCode(3letter)を書き込む
ws2.cell(row=1, column=2).value = '3letterCountryCode'
for i, cc in enumerate(cclist3letter , start=2):
    ws2.cell(row=i, column=2).value = cc


# Excelのコピー作業
for plantfile in os.listdir(folder_path):
    if plantfile.endswith('.xlsx'):
        #フォルダの中のExcel植物ファイルを取得
        workbook = px.load_workbook(os.path.join(folder_path,plantfile ))
        ws = workbook.active
        
        # ファイル名から植物名を抽出する
        plantnames=plantfile.split('.')
        plantName=plantnames[0]
        print(plantName)

        
        ws2 = newWb.active
        ws2.cell(row=1, column=k).value=plantName


        #CountryCode格納用のリストを作成
        newccList=[]       
        # リストにその植物のすべてのCountryCodeを格納
        for row in ws.iter_rows():
            for cell in row:
                if cell.col_idx == 16:
                    newccList.append(cell.value)
                    
        # 格納されたリストの数を確認
        print(len(newccList)) 

        i=2
        for cc in cclist:
            ws2.cell(row=i, column=k).value=newccList.count(cc)
            i = i + 1 

        k=k+1

print("done")

newWb.save('SumCountryCode.xlsx')  

# Excel ファイルを読み込む
df = pd.read_excel('SumCountryCode.xlsx')

# DataFrame を CSV ファイルに出力する
df.to_csv('SumCountryCode.csv', index=False)  